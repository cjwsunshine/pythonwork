"""
Dify自定义工具：JSON转Excel转换器
在Dify工作流中作为工具节点使用
"""

import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
import base64
import io
import os
import tempfile


class JSONToExcelTool:
    """JSON转Excel工具类"""

    def __init__(self):
        self.name = "JSON转Excel转换器"
        self.description = "将JSON格式的测试用例转换为Excel文件"
        self.version = "1.0.0"

    def get_tool_config(self) -> Dict:
        """获取工具配置"""
        return {
            "name": self.name,
            "description": self.description,
            "parameters": {
                "type": "object",
                "properties": {
                    "json_input": {
                        "type": "string",
                        "description": "JSON格式的测试用例数据"
                    },
                    "output_format": {
                        "type": "string",
                        "enum": ["excel", "csv"],
                        "default": "excel",
                        "description": "输出格式"
                    },
                    "include_styling": {
                        "type": "boolean",
                        "default": True,
                        "description": "是否包含Excel样式"
                    },
                    "filename": {
                        "type": "string",
                        "default": "test_cases",
                        "description": "输出文件名（不含扩展名）"
                    }
                },
                "required": ["json_input"]
            }
        }

    def run(self, json_input: str, **kwargs) -> Dict:
        """
        执行转换

        Args:
            json_input: JSON字符串或文件路径
            **kwargs: 其他参数

        Returns:
            包含转换结果的字典
        """
        try:
            # 解析输入参数
            output_format = kwargs.get('output_format', 'excel')
            include_styling = kwargs.get('include_styling', True)
            filename = kwargs.get('filename', 'test_cases')

            # 解析JSON数据
            data = self._parse_json_input(json_input)

            # 提取测试用例
            test_cases = self._extract_test_cases(data)

            if output_format == 'excel':
                # 生成Excel文件
                result = self._generate_excel(test_cases, filename, include_styling)
            else:
                # 生成CSV文件
                result = self._generate_csv(test_cases, filename)

            return {
                "success": True,
                "data": result,
                "message": f"成功转换 {len(test_cases)} 个测试用例",
                "metadata": {
                    "test_case_count": len(test_cases),
                    "output_format": output_format,
                    "generated_at": datetime.now().isoformat()
                }
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": "转换失败"
            }

    def _parse_json_input(self, json_input: str) -> Any:
        """解析JSON输入"""
        # 检查是否是文件路径
        if os.path.exists(json_input):
            with open(json_input, 'r', encoding='utf-8') as f:
                return json.load(f)

        # 否则作为JSON字符串解析
        return json.loads(json_input)

    def _extract_test_cases(self, data: Any) -> List[Dict]:
        """从数据中提取测试用例"""
        test_cases = []

        if isinstance(data, list):
            # 直接是测试用例列表
            test_cases = data
        elif isinstance(data, dict):
            # 查找常见的测试用例字段
            possible_keys = ['test_cases', 'cases', 'testCases', 'testcases', 'items']
            for key in possible_keys:
                if key in data and isinstance(data[key], list):
                    test_cases = data[key]
                    break

            # 如果没有找到，检查其他结构
            if not test_cases:
                # 深度搜索测试用例
                test_cases = self._deep_search_test_cases(data)

        return test_cases

    def _deep_search_test_cases(self, data: Any) -> List[Dict]:
        """深度搜索测试用例"""
        results = []

        if isinstance(data, dict):
            # 检查当前对象是否是测试用例
            if self._is_test_case(data):
                results.append(data)
            else:
                # 递归搜索
                for value in data.values():
                    results.extend(self._deep_search_test_cases(value))

        elif isinstance(data, list):
            for item in data:
                results.extend(self._deep_search_test_cases(item))

        return results

    def _is_test_case(self, obj: Dict) -> bool:
        """判断是否为测试用例"""
        # 测试用例通常有这些字段
        test_case_indicators = ['id', 'case_id', 'title', 'name', 'steps']
        return any(indicator in obj for indicator in test_case_indicators)

    def _generate_excel(self, test_cases: List[Dict], filename: str, include_styling: bool) -> Dict:
        """生成Excel文件"""
        # 准备数据
        rows = []
        for case in test_cases:
            row = self._prepare_test_case_row(case)
            rows.append(row)

        # 创建DataFrame
        df = pd.DataFrame(rows)

        # 创建Excel写入器
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='测试用例')

            if include_styling:
                self._apply_excel_styling(writer)

        # 获取文件内容
        excel_data = output.getvalue()

        # 返回结果
        return {
            "file_content": base64.b64encode(excel_data).decode('utf-8'),
            "filename": f"{filename}.xlsx",
            "file_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "download_url": None,  # Dify可以处理文件返回
            "preview": self._generate_preview(df)
        }

    def _generate_csv(self, test_cases: List[Dict], filename: str) -> Dict:
        """生成CSV文件"""
        # 准备数据
        rows = []
        for case in test_cases:
            row = self._prepare_test_case_row(case)
            rows.append(row)

        # 创建DataFrame
        df = pd.DataFrame(rows)

        # 生成CSV
        csv_data = df.to_csv(index=False, encoding='utf-8-sig')

        return {
            "file_content": csv_data,
            "filename": f"{filename}.csv",
            "file_type": "text/csv",
            "preview": df.head().to_string()
        }

    def _prepare_test_case_row(self, test_case: Dict) -> Dict:
        """准备测试用例行数据"""
        # 基本字段
        row = {
            '用例ID': test_case.get('id', test_case.get('case_id', '')),
            '用例名称': test_case.get('title', test_case.get('name', '')),
            '模块': test_case.get('module', test_case.get('category', '')),
            '优先级': test_case.get('priority', ''),
            '前置条件': test_case.get('preconditions', ''),
            '测试步骤': self._format_steps(test_case.get('steps', [])),
            '预期结果': self._format_expected(test_case.get('expected', '')),
            '测试数据': self._format_test_data(test_case.get('test_data', {})),
            '状态': test_case.get('status', '未执行'),
            '创建人': test_case.get('author', ''),
            '创建时间': test_case.get('created_date', ''),
            '备注': test_case.get('remarks', '')
        }

        return row

    def _format_steps(self, steps: Any) -> str:
        """格式化测试步骤"""
        if isinstance(steps, list):
            formatted = []
            for i, step in enumerate(steps, 1):
                if isinstance(step, dict):
                    action = step.get('action', step.get('step', ''))
                    formatted.append(f"{i}. {action}")
                elif isinstance(step, str):
                    formatted.append(f"{i}. {step}")
            return '\n'.join(formatted)
        elif isinstance(steps, str):
            return steps
        return ''

    def _format_expected(self, expected: Any) -> str:
        """格式化预期结果"""
        if isinstance(expected, list):
            return '\n'.join([f"{i + 1}. {exp}" for i, exp in enumerate(expected)])
        elif isinstance(expected, str):
            return expected
        return ''

    def _format_test_data(self, test_data: Any) -> str:
        """格式化测试数据"""
        if isinstance(test_data, dict):
            return '\n'.join([f"{k}: {v}" for k, v in test_data.items()])
        elif isinstance(test_data, str):
            return test_data
        return ''

    def _apply_excel_styling(self, writer):
        """应用Excel样式"""
        workbook = writer.book
        worksheet = writer.sheets['测试用例']

        # 定义样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="4F81BD",
            end_color="4F81BD",
            fill_type="solid"
        )

        # 应用表头样式
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # 启用文本换行
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _generate_preview(self, df: pd.DataFrame) -> str:
        """生成数据预览"""
        preview_lines = []
        preview_lines.append(f"数据概览:")
        preview_lines.append(f"总行数: {len(df)}")
        preview_lines.append(f"总列数: {len(df.columns)}")
        preview_lines.append("\n字段列表:")
        for col in df.columns:
            preview_lines.append(f"  - {col}")

        if len(df) > 0:
            preview_lines.append("\n前3行数据:")
            for i in range(min(3, len(df))):
                preview_lines.append(f"行 {i + 1}: {dict(df.iloc[i])}")

        return '\n'.join(preview_lines)


# Dify工具注册函数
def get_tool() -> Dict:
    """获取工具定义（Dify要求的标准接口）"""
    tool_instance = JSONToExcelTool()

    return {
        "name": tool_instance.name,
        "description": tool_instance.description,
        "parameters": tool_instance.get_tool_config()["parameters"],
        "execute": tool_instance.run
    }


# 测试函数
if __name__ == "__main__":
    # 测试示例
    test_json = {
        "test_cases": [
            {
                "id": "TC001",
                "title": "用户登录测试",
                "module": "认证模块",
                "priority": "高",
                "steps": [
                    {"step": 1, "action": "打开登录页面", "expected": "页面正常显示"},
                    {"step": 2, "action": "输入用户名密码", "expected": "输入成功"}
                ],
                "status": "未执行"
            }
        ]
    }

    tool = JSONToExcelTool()
    result = tool.run(json.dumps(test_json))
    print("转换结果:", result.get("success"))
    print("消息:", result.get("message"))